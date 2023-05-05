# -*- coding: utf-8 -*-
import openpyxl
import sys
import os


EXCEL_SHEETS_DIR_PATH = "./excelsheets"
EXCEL_SHEET_TARGET_PATH = "./SheetTarget.xlsx"
COLUMNS = ["CO2 Emission"]


def validate_sheet(ws):
    # Print general information
    max_row = ws.max_row
    max_column = ws.max_column
    print(f"Sheet name: {ws.title}")
    print(f"Total number of rows: {str(max_row)} number of columns: {str(max_column)}")

    # General sanity checks
    if max_row < 2:
        sys.exit("No data found in the sheet")
    if max_column != len(COLUMNS):
        sys.exit(f"Expected to find {str(COLUMNS)} columns but found {max_column} column(s)")
    for i in range(1, max_column + 1):
        if ws.cell(row=1, column=i).value != COLUMNS[i - 1]:
            sys.exit(f"Expected to find column {COLUMNS[i - 1]} but found {ws.cell(row=1, column=i).value}")


def append_to_sheet(ws_target, ws_source):
    max_row = ws_source.max_row
    for i in range(2, max_row):
        val = ws_source.cell(row=i, column=1).value
        ws_target.append([val])


def aggregate_sheets(dir_path):
    wb_target = openpyxl.Workbook()
    ws_target = wb_target.active
    ws_target.append(COLUMNS)

    for filename in os.listdir(dir_path):
        if filename.endswith(".xlsx"):
            f = os.path.join(dir_path, filename)
            wb_tmp = openpyxl.load_workbook(f)
            ws_tmp = wb_tmp.active
            validate_sheet(ws_tmp)
            append_to_sheet(ws_target, ws_tmp)
    
    # May override existing file
    wb_target.save(EXCEL_SHEET_TARGET_PATH)


if __name__ == "__main__":
    aggregate_sheets(EXCEL_SHEETS_DIR_PATH)
