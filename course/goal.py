import os
import shutil
import openpyxl
from collections import defaultdict


RESULT_LOCATION = "/Users/tguehring/Downloads/testtesttest"

SAMPLE_TEMPLATE_EXCEL_SHEET = "/Users/tguehring/Projects/private/python-excel-utilities/course/Region.xlsx"
VALIDATION_EXCEL_SHEET = "/Users/tguehring/Projects/private/python-excel-utilities/course/Validation.xlsx"
REGION_COL_IDX = 1
CRE_ID_COL_IDX = 2


class SampleTemplateRow:
    def __init__(self, region, creid, address, sampling_period, sampling_value) -> None:
        self.region = region
        self.creid = creid
        self.address = address
        self.sampling_period = sampling_period
        self.sampling_value = sampling_value

    def __str__(self):
        return f"Region: {self.region}, CREID: {self.creid}, Address: {self.address}, Sampling Period: {self.sampling_period}, Sampling Value: {self.sampling_value}"


def copy_file_into_dir(source_file_path: str, target_location: str) -> str:
    """
    To copy file into target directory
    """
    return shutil.copy(source_file_path, target_location)


def get_regions_from_excel_sheet(excel_sheet_path: str):
    workbook = openpyxl.load_workbook(excel_sheet_path)
    ws = workbook['Sheet1']
    max_row = ws.max_row
    print(f"Max row: {max_row}")
    
    region_row_map = defaultdict(list)
    # Iterate through all the rows to get the regions
    # Start at row 2 to exclude the header 'Region'
    for i in range(2, max_row + 1):
        region = ws['A' + str(i)].value
        sample_template_row = SampleTemplateRow(
            region,
            ws['B' + str(i)].value,
            ws['C' + str(i)].value,
            ws['D' + str(i)].value,
            ws['E' + str(i)].value
        )

        print(f"Processing row: {sample_template_row}")
        region_row_map[region].append(sample_template_row)

    return region_row_map


def sanitize_excel_sheet(file_path: str, region: str):
    """
    To remove all rows that do not belong to the region
    """
    print("Sanitizing file: " + file_path)

    workbook = openpyxl.load_workbook(file_path)
    ws = workbook['Sheet1']
    max_row = ws.max_row
    # Iterate through all the rows to get the regions
    # Start at row 2 to exclude the header 'Region'
    i = 2
    while i <= max_row:
        row_region = ws.cell(row=i, column=REGION_COL_IDX).value
        print(f"Row {i} has region {row_region}")
        if row_region is None or row_region == "":
            break
        elif row_region != region:
            ws.delete_rows(i)
        else:
            i += 1

    workbook.save(file_path)


def insert_sample_template_row_into_excel_sheet(file_path: str, sample_template_row: SampleTemplateRow):
    """
    Insert sample_template_row in line 4 into target_excel_sheet
    """
    print("Inserting row: " + str(sample_template_row) + " into file: " + file_path)

    workbook = openpyxl.load_workbook(file_path)
    ws = workbook['Sheet1']
    max_row = ws.max_row
    print(f"Max row: {max_row}")

    # Insert row at line 4
    ws.insert_rows(4)
    # Write values into the row
    ws['A4'] = sample_template_row.region
    ws['B4'] = sample_template_row.creid
    ws['C4'] = sample_template_row.address
    ws['D4'] = sample_template_row.sampling_period
    ws['E4'] = sample_template_row.sampling_value

    workbook.save(file_path)


def setup_region_folders(region_row_map, target_excel_sheet: str):
    """
    To create folders for each region and copy the target_excel_sheet into it
    """
    for region in region_row_map:
        region_dir_path = os.path.join(RESULT_LOCATION, region)
        os.makedirs(region_dir_path, exist_ok=True)
        copy_file_into_dir(SAMPLE_TEMPLATE_EXCEL_SHEET, region_dir_path)
        for sample_template_row in region_row_map[region]:
            creid_dir_path = os.path.join(region_dir_path, sample_template_row.creid)
            os.makedirs(creid_dir_path, exist_ok=True)
            print("Created folder: ", creid_dir_path)
            copied_excel_file = copy_file_into_dir(target_excel_sheet, creid_dir_path)
            insert_sample_template_row_into_excel_sheet(copied_excel_file, sample_template_row)


region_row_map = get_regions_from_excel_sheet(SAMPLE_TEMPLATE_EXCEL_SHEET)
setup_region_folders(region_row_map, VALIDATION_EXCEL_SHEET)