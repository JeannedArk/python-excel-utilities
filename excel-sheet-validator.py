# -*- coding: utf-8 -*-
import openpyxl
import sys


EXCEL_SHEET_PATH = "./excelsheets/Sheet1.xlsx"
COLUMNS = ["CO2 Emission"]
CO2_MAX_VAL = 10000000


def validate_sheet(ws):
    # Print general information
    max_row = ws.max_row
    max_column = ws.max_column
    print(f"Total number of rows: {str(max_row)} number of columns: {str(max_column)}")

    # General sanity checks
    if max_row < 2:
        sys.exit("No data found in the sheet")
    if max_column != len(COLUMNS):
        sys.exit(f"Expected to find {str(COLUMNS)} columns but found {max_column} column(s)")
    for i in range(1, max_column + 1):
        if ws.cell(row=1, column=i).value != COLUMNS[i - 1]:
            sys.exit(f"Expected to find column {COLUMNS[i - 1]} but found {ws.cell(row=1, column=i).value}")

    # Data specific sanity checks
    for i in range(2, max_row):
        val = ws.cell(row=i, column=1).value
        if val > CO2_MAX_VAL:
            sys.exit(f"CO2 emission value {val} at row {i} exceeds maximum value of {CO2_MAX_VAL}")
    print(f"Data is valid for sheet: {ws.title}")


if __name__ == "__main__":
    wb = openpyxl.load_workbook(EXCEL_SHEET_PATH)
    ws = wb.active
    validate_sheet(ws)
